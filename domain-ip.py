import os
import openpyxl as xl
from socket import gethostbyname, gaierror

def read_excel(column, row):
    workbook = xl.load_workbook('./out/输出信息.xlsx')
    sheet = workbook['子域信息']
    value = sheet[column + str(row)].value
    workbook.close()
    return value

def ipjx(value):
    try:
        host = gethostbyname(value)  # 域名反解析得到的IP
        print("%s\t%s" % (host, value))
        return host
    except gaierror as err:
        print("域名: %s 未解析到IP.(错误: %s)" % (value, err))

def append_host(host, host_list):
    if host is not None and host not in host_list:
        host_list.append(host)
        return host_list


def data_saver(value_list, host_list):
    file_path = './out/输出信息.xlsx'
    if os.path.exists(file_path):
        wb = xl.load_workbook(file_path)
    else:
        wb = xl.Workbook()
    if 'IP信息' not in wb.sheetnames:
        ws = wb.create_sheet('IP信息')
    else:
        ws = wb['IP信息']
    ws.cell(1, 1).value = "子域名"
    ws.cell(1, 2).value = "IP地址"
    for i, val in enumerate(value_list, start=2):
        ws.cell(i, 1).value = val
    for i, host in enumerate(host_list, start=2):
        ws.cell(i, 2).value = host

    # 设置列宽为 45
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 45

    wb.save(file_path)
    wb.close()

def main():
    host_list = []
    value_list = []
    for column in range(ord('A'), ord('Z')+1):
        for row in range(2, 10):
            value = read_excel(chr(column), row)
            if value is not None:
                print("正在查询 " + value + " 的IP信息")
                host = ipjx(value)
                append_host(host, host_list)
                value_list.append(value)
    print(host_list)
    data_saver(value_list, host_list)

if __name__ == "__main__":
    main()
