#!/usr/bin/python3
# coding:utf-8
# Author:se55i0n
# 目标tcp端口开放扫描及应用端口banner识别

import os
import sys
import socket
import logging
import requests
import dns.resolver
import openpyxl as xl
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from requests.packages.urllib3.exceptions import InsecureRequestWarning
# 线程池
from multiprocessing.dummy import Pool as ThreadPool
from multiprocessing.dummy import Lock

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

logger = logging.getLogger(__name__)

def read_ip(column, row):
    workbook = xl.load_workbook('./out/输出信息.xlsx')
    sheet = workbook['IP信息']
    ip = sheet[column + str(row)].value
    workbook.close()
    return ip


def check_cdn(target):
    # 目标域名cdn检测
    myResolver = dns.resolver.Resolver()
    myResolver.lifetime = myResolver.timeout = 2.0
    try:
        result = [['114.114.114.114'], ['8.8.8.8'], ['223.6.6.6']]
        for i in result:
            myResolver.nameservers = i
            record = myResolver.resolve(target)
            result.append(record[0].address)
        return True if len(set(result)) > 1 else False
    except Exception as e:
        pass
    finally:
        return False


def scan_port(target, port):
    # 端口扫描
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(0.2)
        return True if s.connect_ex((target, port)) == 0 else False
    except Exception as e:
        pass
    finally:
        s.close()

def get_http_banner(url):
    # http/https请求获取banner
    try:
        r = requests.get(url, headers={'UserAgent': UserAgent().random},
                            timeout=2, verify=False, allow_redirects=True)
        soup = BeautifulSoup(r.content, 'lxml')
        return soup.title.text.strip('\n').strip()
    except Exception as e:
        pass

def get_socket_info(target, port):
    # socket获取banner
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(0.2)
        s.connect((target, port))
        s.send(b'HELLO\r\n')
        return s.recv(1024).split(b'\r\n')[0].strip(b'\r\n')
    except Exception as e:
        pass
    finally:
        s.close()

def run(target, port, mutex, result):
    try:
        if scan_port(target, port):
            banner = get_http_banner(f'http://{target}:{port}')
            mutex.acquire()
            if banner:
                print('[+] {} ---- open   {}'.format(
                    str(port).rjust(6), banner[:18]))
            else:
                banner = get_http_banner(f'https://{target}:{port}')
                if banner:
                    print('[+] {} ---- open   {}'.format(
                        str(port).rjust(6), banner[:18]))
                else:
                    banner = get_socket_info(target, port)
                    if banner:
                        print('[+] {} ---- open   {}'.format(
                            str(port).rjust(6), banner[:18]))
                    else:
                        print('[+] {} ---- open'.format(
                            str(port).rjust(6)))
            result.append(port)
            mutex.release()
    except Exception as e:
        pass

def start_scan(target):
    try:
        print('[-] 正在扫描地址: {} '.format(socket.gethostbyname(target)))
        # 线程数
        pool = ThreadPool(processes=100)
        # get传递超时时间，用于捕捉ctrl+c
        ports = range(1, 65536)
        mutex = Lock()
        result = []
        pool.map_async(lambda port: run(target, port, mutex, result), ports).get(0xffff)
        pool.close()
        pool.join()
        return result
    except Exception as e:
        print(e)
        return []
    except KeyboardInterrupt:
        print('\n[-] 用户终止扫描...')
        sys.exit(1)

def data_saver(row, ip, ports):
    file_path = './out/输出信息.xlsx'
    if os.path.exists(file_path):
        wb = xl.load_workbook(file_path)
    else:
        wb = xl.Workbook()
    if '端口信息' not in wb.sheetnames:
        ws = wb.create_sheet('端口信息')
    else:
        ws = wb['端口信息']
    ws.cell(1, row - 1).value = ip + "开放端口"
    for i, val in enumerate(ports, start=2):
        ws.cell(i, row - 1).value = val

    ws.column_dimensions[chr((row - 1) + 64)].width = 45

    wb.save(file_path)
    wb.close()


def main():
    for column in range(ord('B'), ord('Z')+1):
        for row in range(2, 100):
            ip = read_ip(chr(column), row)
            if ip is None:  # 检查IP地址是否为空
                break  # 停止循环
            else:
                if not check_cdn(ip):
                    ports = start_scan(ip)
                    print("扫描结束，准备写入")
                    data_saver(row, ip, ports)
                else:
                    print('[-] 检测到CDN技术,停止扫描.')
        else:
            continue  # 内循环正常结束，继续下一个外循环
        break  # 内循环被break跳出，停止外循环




if __name__ == "__main__":
    main()
