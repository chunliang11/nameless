import re
import requests
import openpyxl

blacklist = ["黑名单词汇1", "黑名单词汇2", "黑名单词汇3"]


def read_web():
    file_path = './out/输出信息.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    max_column = sheet.max_column
    max_row = sheet.max_row
    url_list = []
    for column_index in range(1, max_column + 1):
        # 获取第一行的值作为第一个参数
        ip = sheet.cell(row=1, column=column_index).value
        # 从第二行开始循环读取每一行，直到遇到空值
        row_index = 2
        while row_index <= max_row:
            port = sheet.cell(row=row_index, column=column_index).value
            if port is None:
                break
            url = "http://" + ip + ":" + str(port)
            url_list.append(url)
            row_index += 1
    return url_list


def getTitle(url_list):
    for i, url in enumerate(url_list):
        print("当前正在处理第(%d)条数据, 共计(%s)条数据." % (i + 1, len(url_list)))
        print(url)
        html_list = getHTML(url.strip())
        getTitleFromHTML(html_list)



def getTitleFromHTML(html_list):
    title = None
    html = html_list[0]
    code = html_list[1]
    status = html_list[2]
    url = html_list[3]

    if code == 200:
        pattern = re.compile('<title>(.*?)</title>', re.I)
        if pattern.search(html):
            title = pattern.search(html).group(1)
        else:
            pattern = re.compile("window.location='\./(.*?)'", re.I)
            if pattern.search(html):
                url = "%s/%s" % (url, pattern.search(html).group(1))
                return getTitleFromHTML(getHTML(url))
            else:
                status = '非正常响应页面'

    return [title, code, status]


def getHTML(url):
    html = None
    code = None
    status = '请求正常'
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:52.0) Gecko/20100101 Firefox/52.0"
    }

    if '://' not in url:
        url = "http://%s" % url if "443" not in url else "https://%s" % url

    try:
        response = requests.get(url, headers=headers, timeout=3)
        response.encoding = response.apparent_encoding
        html = response.text.replace('\t', '').replace('\r', '').replace('\n', '')
        code = response.status_code
    except requests.exceptions.RequestException as err:
        status = str(err)

    # 如果使用 http 协议时出现非正常响应，尝试使用 https 协议进行请求
    if status != '请求正常' and url.startswith("http://"):
        url_https = url.replace("http://", "https://")
        try:
            response = requests.get(url_https, headers=headers, timeout=3)
            response.encoding = response.apparent_encoding
            html = response.text.replace('\t', '').replace('\r', '').replace('\n', '')
            code = response.status_code
            status = '请求正常'
            url = url_https
        except requests.exceptions.RequestException as err:
            pass

    return [html, code, status, url]


def save_web(url_list):
    file_path = './out/输出信息.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.create_sheet(title='web信息')
    sheet.cell(row=1, column=1, value='URL')
    sheet.cell(row=1, column=2, value='标题')

    row_index = 2
    normal_responses = []
    error_responses = []

    for url in url_list:
        html_list = getHTML(url.strip())
        title, code, status = getTitleFromHTML(html_list)

        if code == 200 and title not in blacklist:
            normal_responses.append((url, title))
        else:
            error_responses.append((url, code, title, status))
    # 写入正常响应
    for url, title in normal_responses:
        sheet.cell(row=row_index, column=1, value=url)
        sheet.cell(row=row_index, column=2, value=title)
        row_index += 1
    # 添加空行
    row_index += 1
    # 写入非正常响应
    for url, code, title, status in error_responses:
        sheet.cell(row=row_index, column=1, value=url)
        sheet.cell(row=row_index, column=2, value="%s: %s: %s" % (code, title, status))
        row_index += 1

    workbook.save(file_path)




def main():
    url_list = read_web()
    getTitle(url_list)
    save_web(url_list)


if __name__ == "__main__":
    main()
