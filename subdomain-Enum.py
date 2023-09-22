#!/usr/local/bin/python
import requests
import openpyxl
import traceback
import json
import os
import openpyxl as xl
from openpyxl.styles import Alignment

def read_domain(start_row):
    try:
        workbook = openpyxl.load_workbook('./out/输出信息.xlsx')
        sheet = workbook['备案信息']
        column = 'B'
        parameter = sheet[column + str(start_row)].value
        while parameter is not None:
            print("正在查询"+parameter+"的子域名")
            return parameter
        workbook.close()
    except:
        return -1


def get_sub_domains(domain, x=0):
    arr_key = ["XzkUwrmIwQfZwad3rSwTVitlB1BkFXOi", "QkNPITjv96R5jlYtJaDI4aLsXd2xBpzh", "m_T4d1SKYxnOs1c80HVtbYS5mtI31aEc", "FVneqp2-u3derLyiOc_cn7g9kc_sdPu_", "xhiqba77g3b8QwidakbNjVXRrFJiGNC0", "UBwe7CNGywbulPnoUjh9RJnnjejY9DWN", "MX2bd7Cfl0H2ocYh3Q2x8bDAdxqkNv_a", "fQ7BgfhC2lPmwfZb8bFdNmy2p2DcObXe", "O_jgoFks_Qve7z6sNaYec45YeMstJHyw", "00Euy2uEZwDD8F262Bah5iuE6369HtHU"]
    if x >= len(arr_key):
        return -1  # 所有的键都失效了

    keyapi = arr_key[x]
    try:
        url = "https://api.securitytrails.com/v1/domain/"+domain+"/subdomains"
        querystring = {"children_only":"true"}
        headers = {
            'accept': "application/json",
            'apikey': keyapi
        }
        response = requests.request("GET", url, headers=headers, params=querystring)
        result_json = json.loads(response.text)
        sub_domains = [i+'.'+domain for i in result_json['subdomains']]
        return sub_domains
    except KeyError:
        print(f"key[{x}]失效")
    except Exception as e:
        traceback.print_exc()

    if x < len(arr_key) - 1:
        return get_sub_domains(domain, x=x+1)  # 使用下一个键

    return -1  # 所有的键都失效了



def data_saver(sub_domains):
    """
    打印最终结果，并保存数据至Excel表格，同时调整表格格式。
    """
    total_row = len(sub_domains)
    if total_row == 1:
        total_row = 0
    elif total_row == 0:
        return print("没有解析到子域\n")
    print(f"查询结果如下:\n\n{sub_domains}\n")
    file_path = './out/输出信息.xlsx'
    # 存在对应文件，则读取表格追加写入，不存在则创建，并设置表格的标题、列宽、冻结窗格、文字布局等格式
    if os.path.exists(file_path):
        wb = xl.load_workbook(file_path)
    else:
        wb = xl.Workbook()
    if '子域信息' not in wb.sheetnames:
        ws = wb.create_sheet('子域信息')
        title_list = ['子域信息']
        for i in range(len(title_list)):
            ws.cell(1, i + 1).value = title_list[i]
        col_width = {'A': 45, 'B': 40, 'C': 22, 'D': 24, 'E': 9, 'F': 15, 'G': 13, 'H': 21}
        for k, v in col_width.items():
            ws.column_dimensions[k].width = v
        ws.freeze_panes = 'A2'
    else:
        ws = wb['子域信息']
    max_row = ws.max_row
    start = max_row + 1
    total_row = total_row + start
    after_title = 0
    # 写入查询数据
    ws.append(sub_domains)

    # 垂直居中
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            ws.cell(row + 1, col + 1).alignment = Alignment(horizontal='center', vertical='center')
    try:
        wb.save(file_path)
    except PermissionError:
        print("** 备案信息登记表格已打开，无法写入文件。如需写入，请关闭文件后重新执行！ **\n")
        return -1
    print(f"查询结果保存在：{file_path}\n")
    return 'OK'



   

def main():
    start_row = 2
    while start_row < 500:
        domain = read_domain(start_row)
        if domain is None:
            break
        start_row += 1
        sub_domains = get_sub_domains(domain, x=0)
        if sub_domains != -1:
            data_saver(sub_domains)
        else:
            raise ValueError("获取key失败，请检查key是否生效，或超过最大使用次数！")
          


if __name__ == '__main__':
    main()